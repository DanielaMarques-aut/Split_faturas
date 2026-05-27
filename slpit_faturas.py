
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

# ── CONFIGURAÇÃO ──────────────────────────────────────────────────────────────
# Ajuste os nomes dos ficheiros e a lista de PMs para dividir por cliente conforme necessário.
FATURAS_FILE = "Módulo de Faturas.xlsx"
PMO_FILE     = "Listagem PMO Abril.xlsx"
OUTPUT_DIR   = "Faturas"
SPLIT_BY_CLIENT_PMS: list[str] = ["JOSE TORRES", "PEDRO PINELA"]  # em maiúsculas
# ─────────────────────────────────────────────────────────────────────────────

GREEN      = "C6EFCE"
RED        = "FFC7CE"
HEADER_BG  = "1F4E79"
HEADER_FG  = "FFFFFF"

def style_headers(ws, source_ws):
    """Copia o estilo dos cabeçalhos da sheet original."""
    for col in range(1, source_ws.max_column + 1):
        src = source_ws.cell(row=1, column=col)
        dst = ws.cell(row=1, column=col, value=src.value)
        dst.font = Font(
            bold=src.font.bold,
            color=src.font.color.rgb if src.font.color and src.font.color.type == "rgb" else HEADER_FG,
            name=src.font.name or "Arial",
            size=src.font.size or 10
        )
        if src.fill and src.fill.fgColor and src.fill.fgColor.type == "rgb":
            dst.fill = PatternFill("solid", start_color=src.fill.fgColor.rgb)
        else:
            dst.fill = PatternFill("solid", start_color=HEADER_BG)
        dst.alignment = Alignment(
            horizontal=src.alignment.horizontal or "center",
            vertical=src.alignment.vertical or "center"
        )
        ws.column_dimensions[get_column_letter(col)].width = (
            source_ws.column_dimensions[get_column_letter(col)].width or 14
        )
    ws.row_dimensions[1].height = source_ws.row_dimensions[1].height or 20

def apply_row_color(ws, row_num, pago_col, vencida_col, n_cols):
    pago    = str(ws.cell(row=row_num, column=pago_col).value or "").strip().upper()
    vencida = str(ws.cell(row=row_num, column=vencida_col).value or "").strip().upper()

    if pago == "PAGO" and vencida == "NÃO":
        color = GREEN
    elif pago == "NÃO PAGO" or vencida == "SIM":
        color = RED
    else:
        return

    fill = PatternFill("solid", start_color=color)
    for col in range(1, n_cols + 1):
        ws.cell(row=row_num, column=col).fill = fill

def match_wbs(wbs, contract):
    wbs = str(wbs).strip()
    contract = str(contract).strip()
    if not wbs or not contract:
        return False
    prefix = wbs.split(".")[0]
    return prefix.startswith(contract)

def load_faturas(path):
    wb = load_workbook(path)
    main_ws = wb.worksheets[0]
    print(f"  → '{main_ws.title}' sheet loaded with {main_ws.max_row - 1} data rows.")
    all_rows = []
    headers = None
    pago_col = vencida_col = wbs_col = None

    # Read headers from first sheet
    for col in range(1, main_ws.max_column + 1):
        val = str(main_ws.cell(row=1, column=col).value or "").strip().upper()
        if val:
            if headers is None:
                headers = {}
            headers[col] = main_ws.cell(row=1, column=col).value

    # Find key columns
    for col, name in headers.items():
        n = str(name).strip().upper()
        if "WBS" in n:
             wbs_col = col
        if n == "PAGO":
            pago_col = col
        if n == "VENCIDA":
            vencida_col = col

    # Read all data rows from all sheets
    
    max_row =main_ws.max_row 
    print(f"  → Processing '{main_ws.title}' with {max_row - 1} data rows...")
    if max_row < 2:
        print("  → No data rows to process.")
        return wb, main_ws, headers, all_rows, wbs_col, pago_col, vencida_col
    for row in range(2, max_row + 1):
        if all(main_ws.cell(row=row, column=c).value is None for c in range(1, main_ws.max_column + 1)):
            continue
        all_rows.append([main_ws.cell(row=row, column=c).value for c in range(1, main_ws.max_column + 1)])
    print (f"  → Total de {len(all_rows)} faturas carregadas.")
    return wb, main_ws, headers, all_rows, wbs_col, pago_col, vencida_col
    

def load_pmo(path):
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    contrato_col = chefe_col = cliente_col = None

    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or "").strip().upper()
        if "CONTRATO" in h:
             contrato_col = col
        if "CHEFE" in h:     
            chefe_col = col
        if "CLIENTE" in h:  
            cliente_col = col

    pm_contracts = {}   # pm -> {cliente -> [contratos]}
    last_row = ws.max_row

    for row in range(2, last_row + 1):
        pm      = str(ws.cell(row=row, column=chefe_col).value or "").strip()
        contrato = str(ws.cell(row=row, column=contrato_col).value or "").strip()
        cliente  = str(ws.cell(row=row, column=cliente_col).value or "").strip()
        if not pm or not contrato:
            continue
        if pm not in pm_contracts:
            pm_contracts[pm] = {}
        if cliente not in pm_contracts[pm]:
            pm_contracts[pm][cliente] = []
        pm_contracts[pm][cliente].append(contrato)

    return pm_contracts

def write_pm_file(path, rows, headers, source_ws, pago_col, vencida_col):
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Faturas"

    style_headers(ws_out, source_ws)

    n_cols = len(headers)
    for i, row in enumerate(rows, start=2):
        for col in range(1, n_cols + 1):
            ws_out.cell(row=i, column=col, value=row[col - 1])
        apply_row_color(ws_out, i, pago_col, vencida_col, n_cols)

    wb_out.save(path)

def clean_name(name):
    invalid = r'\/:*?"<>|'
    result = "".join("_" if c in invalid else c for c in name)
    return result.strip()[:50]

def main():
    for f in [FATURAS_FILE, PMO_FILE]:
        if not Path(f).exists():
            print(f"ERRO: Ficheiro não encontrado: {f}")
            sys.exit(1)

    out = Path(OUTPUT_DIR)
    out.mkdir(exist_ok=True)

    print("A ler Módulo de Faturas...")
    wb, main_ws, headers, all_rows, wbs_col, pago_col, vencida_col = load_faturas(FATURAS_FILE)
    print("A ler Listagem PMO...")
    pm_contracts = load_pmo(PMO_FILE)

    count = 0
    all_contratos_flat = [c for clientes in pm_contracts.values() for cs in clientes.values() for c in cs]

    for pm, clientes in pm_contracts.items():
        pm_upper = pm.strip().upper()
        split_by_client = pm_upper in SPLIT_BY_CLIENT_PMS

        if split_by_client:
            for cliente, contratos in clientes.items():
                rows_pm = [r for r in all_rows if any(match_wbs(r[wbs_col - 1], c) for c in contratos)]
                if not rows_pm:
                    continue
                fname = out / f"{clean_name(pm)}_{clean_name(cliente)}.xlsx"
                write_pm_file(fname, rows_pm, headers, main_ws, pago_col, vencida_col)
                print(f"  ✓ {pm} / {cliente}: {len(rows_pm)} faturas → {fname.name}")
                count += 1
        else:
            contratos = [c for cs in clientes.values() for c in cs]
            rows_pm = [r for r in all_rows if any(match_wbs(r[wbs_col - 1], c) for c in contratos)]
            if not rows_pm:
                continue
            fname = out / f"{clean_name(pm)}.xlsx"
            write_pm_file(fname, rows_pm, headers, main_ws, pago_col, vencida_col)
            print(f"  ✓ {pm}: {len(rows_pm)} faturas → {fname.name}")
            count += 1

    # Faturas sem PM
    sem_pm = [r for r in all_rows if not any(match_wbs(r[wbs_col - 1], c) for c in all_contratos_flat)]
    print(f"\nFaturas sem PM atribuído: {len(sem_pm)}")
    print("  Detalhes:")
    for r in sem_pm:
        wbs = r[wbs_col - 1] if wbs_col else "N/A"
        print(f"    - WBS: {wbs}")
    if sem_pm:
        fname = out / "SEM_PM_ATRIBUIDO.xlsx"
        write_pm_file(fname, sem_pm, headers, main_ws, pago_col, vencida_col)
        print(f"\n  ⚠ {len(sem_pm)} faturas sem PM → SEM_PM_ATRIBUIDO.xlsx")

    print(f"\nConcluído. {count} ficheiros gerados em: {out.resolve()}")

if __name__ == "__main__":
    main()
